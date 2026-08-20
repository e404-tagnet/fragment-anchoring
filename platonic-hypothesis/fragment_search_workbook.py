#!/usr/bin/env python3
"""Create a workbook for reconstructing documents from known text fragments.

Install dependency:
    pip install openpyxl

Run:
    python fragment_search_workbook.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule

OUTPUT = Path("fragment_search_workbook.xlsx")


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ws.columns:
        width = min(max(max(len(str(c.value or "")) for c in column) + 2, 12), 42)
        ws.column_dimensions[column[0].column_letter].width = width


def add_fragments(ws):
    ws.append(["id", "text", "type", "confidence", "known_order", "approx_gap_after", "notes"])
    rows = [
        [1, "nuclear fission", "exact_string", 1.0, "unknown", "", "Example seed"],
        [2, "temporal displacement engine", "exact_string", 1.0, "unknown", "", "Example seed"],
        [3, "", "isolated_word", 0.5, "unknown", "", "Add a distinctive word"],
    ]
    for row in rows:
        ws.append(row)
    validation = DataValidation(type="list", formula1='"exact_string,isolated_word,concept,unknown"')
    ws.add_data_validation(validation)
    validation.add("C2:C1000")
    ws.conditional_formatting.add("D2:D1000", ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))


def add_hypotheses(ws):
    ws.append(["id", "source_fragment_id", "predicted_phrase", "confidence", "reason", "used_in_query", "verified"])
    ws.append([1, 1, "chain reaction", 0.5, "Likely related phrase", "no", "no"])
    ws.append([2, 1, "reactor core", 0.4, "Likely related phrase", "no", "no"])
    ws.append([3, 2, "spacetime", 0.3, "Possible conceptual relation", "no", "no"])
    validation = DataValidation(type="list", formula1='"yes,no"')
    ws.add_data_validation(validation)
    validation.add("F2:G1000")
    ws.conditional_formatting.add("D2:D1000", ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))


def add_queries(ws):
    ws.append(["id", "query", "original_fragments", "hypotheses", "order", "gap_limit_chars", "status", "result_reference", "notes"])
    ws.append([1, '"nuclear fission" "chain reaction"', "1", "1", "A before B", 500, "not_run", "", "Replace with actual search query"])
    ws.append([2, '"temporal displacement engine" "spacetime"', "2", "3", "A before B", 500, "not_run", "", "Replace with actual search query"])
    validation = DataValidation(type="list", formula1='"not_run,queued,run,rejected,promising"')
    ws.add_data_validation(validation)
    validation.add("G2:G1000")


def add_candidates(ws):
    ws.append(["id", "query_id", "candidate_text_or_excerpt", "exact_matches", "distinctive_matches", "order_score", "proximity_score", "coherence_score", "topic_score", "total_score", "decision", "notes"])
    ws.append([1, 1, "", 0, 0, 0, 0, 0, 0, "=SUM(D2:I2)", "unreviewed", "Paste returned text or excerpt here"])
    ws.append([2, 2, "", 0, 0, 0, 0, 0, 0, "=SUM(D3:I3)", "unreviewed", "Paste returned text or excerpt here"])
    validation = DataValidation(type="list", formula1='"unreviewed,promising,rejected,verified"')
    ws.add_data_validation(validation)
    validation.add("K2:K1000")
    ws.conditional_formatting.add("J2:J1000", ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B"))


def add_instructions(ws):
    ws.append(["Step", "Action"])
    instructions = [
        [1, "Enter only text you actually know in Fragments."],
        [2, "Mark isolated words separately from exact consecutive strings."],
        [3, "Use an LLM to propose possible related phrases, but record them as hypotheses."],
        [4, "Create queries using two known fragments plus one hypothesis."],
        [5, "Record every query and its result reference."],
        [6, "Score candidate results without treating LLM guesses as evidence."],
        [7, "Use promising candidates as context for the next expansion round."],
        [8, "Only mark a candidate verified when independent evidence supports it."],
    ]
    for row in instructions:
        ws.append(row)


def build_workbook():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheets = [
        ("Instructions", add_instructions),
        ("Fragments", add_fragments),
        ("Hypotheses", add_hypotheses),
        ("Queries", add_queries),
        ("Candidates", add_candidates),
    ]

    for name, builder in sheets:
        ws = wb.create_sheet(name)
        builder(ws)
        style_sheet(ws)

    wb["Instructions"].column_dimensions["A"].width = 12
    wb["Instructions"].column_dimensions["B"].width = 90
    wb.save(OUTPUT)
    print(f"Created {OUTPUT.resolve()}")


if __name__ == "__main__":
    build_workbook()
